import pdfplumber
import tabula
from openpyxl.utils import get_column_letter
from werkzeug.utils import secure_filename
import os
from flask import Flask, render_template, request, redirect, url_for, send_file
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
import pandas as pd

app = Flask(__name__, template_folder='templates')

UPLOAD_FOLDER = 'uploads'
ALLOWED_EXTENSIONS = {'pdf'}
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER

OUTPUT_FOLDER = os.path.join(os.path.dirname(os.path.abspath(__file__)), 'output')
os.makedirs(OUTPUT_FOLDER, exist_ok=True)
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ---------------------------------------------------------------------------
# EXTRACTION: Tables Only (tabula-py)
# ---------------------------------------------------------------------------

def extract_tables_from_pdf(pdf_path):
    """
    Extract tables using tabula-py.
    Try lattice mode first (for bordered tables), then stream mode as fallback.
    """
    tables = []
    try:
        tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True, lattice=True)
    except Exception:
        pass

    # If lattice found nothing useful, try stream mode
    if not tables or all(t.empty for t in tables):
        try:
            tables = tabula.read_pdf(pdf_path, pages='all', multiple_tables=True, stream=True)
        except Exception as e:
            print(f"Error reading tables: {e}")
            return []

    return [t for t in tables if not t.empty] if tables else []


# ---------------------------------------------------------------------------
# EXTRACTION: All Text + Tables (pdfplumber)
# ---------------------------------------------------------------------------

def extract_pdf_content(pdf_path):
    """
    Use pdfplumber to extract tables with proper column structure.
    For each page, extract tables and any text outside the table areas.
    Returns a list of (non_table_text_lines, tables) per page.
    """
    pages_data = []

    with pdfplumber.open(pdf_path) as pdf:
        for page in pdf.pages:
            # 1. Extract tables as structured data
            table_settings = {
                "vertical_strategy": "lines",
                "horizontal_strategy": "lines",
                "snap_tolerance": 5,
            }
            raw_tables = page.extract_tables(table_settings) or []

            # If lines-based extraction found nothing, try text-based
            if not raw_tables:
                table_settings_fallback = {
                    "vertical_strategy": "text",
                    "horizontal_strategy": "text",
                }
                raw_tables = page.extract_tables(table_settings_fallback) or []

            # 2. Find table bounding boxes to exclude from text extraction
            table_bboxes = []
            found_tables = page.find_tables(table_settings) if raw_tables else []
            for t in found_tables:
                table_bboxes.append(t.bbox)

            # 3. Extract text outside the tables
            non_table_lines = []
            if table_bboxes:
                # Crop page to exclude table regions and extract remaining text
                cropped_page = page
                for bbox in table_bboxes:
                    cropped_page = cropped_page.outside_bbox(bbox)
                remaining_text = cropped_page.extract_text() or ""
                non_table_lines = [l.strip() for l in remaining_text.split('\n') if l.strip()]
            else:
                # No tables found, get all text
                text = page.extract_text() or ""
                non_table_lines = [l.strip() for l in text.split('\n') if l.strip()]

            pages_data.append((non_table_lines, raw_tables))

    return pages_data


def create_excel(pages_data):
    """
    Create an Excel workbook from extracted page data.
    1. Merge all tables across all pages (combining continuation tables).
    2. Try to re-attach non-table text as an extra column when the line
       count matches the table data row count (e.g. PL(B) No column that
       falls outside the detected table bounding box).
    3. Write everything to a single Excel sheet.
    """
    # --- Step 1: Merge tables across pages, and collect text per-page ---
    # We process page-by-page, building merged tables and associating
    # any "extra column" text lines with the rows they belong to.
    merged_tables = []      # list of list-of-rows (each row is a list of cell values)
    extra_col_header = None
    extra_col_values = []   # flat list of extra column values aligned with data rows

    for non_table_lines, raw_tables in pages_data:
        for table in raw_tables:
            if not table or len(table) == 0:
                continue

            num_cols = len(table[0])

            if merged_tables:
                last = merged_tables[-1]
                last_cols = len(last[0])
                if num_cols == last_cols:
                    # Same structure → check for repeated header
                    first_row_str = ' '.join(str(c or '').strip().lower() for c in table[0])
                    header_str = ' '.join(str(c or '').strip().lower() for c in last[0])
                    if first_row_str == header_str:
                        last.extend(table[1:])  # skip repeated header
                    else:
                        last.extend(table)
                else:
                    # Different table structure
                    merged_tables.append(list(table))
                continue

            # First table
            merged_tables.append(list(table))

        # Handle non-table text lines:
        # If the number of text lines matches the number of data rows
        # on this page's table (excluding the header row if it was the
        # first page), treat them as an extra column.
        if non_table_lines and raw_tables:
            data_rows_on_page = sum(len(t) for t in raw_tables)
            # Check if first table on this page had a header that was kept
            # (first ever page) or skipped (continuation page)
            first_table = raw_tables[0]
            if first_table:
                first_row_str = ' '.join(str(c or '').strip().lower() for c in first_table[0])
                # Check if the header line is in the text lines
                header_candidates = [l for l in non_table_lines
                                     if not any(c.isdigit() for c in l.replace(' ', ''))]

                # If text line count == data row count (table rows minus header),
                # or text line count == total table rows, treat as extra column
                if len(non_table_lines) == data_rows_on_page:
                    # First line might be a header for the extra column
                    if extra_col_header is None:
                        extra_col_header = non_table_lines[0]
                        extra_col_values.extend(non_table_lines[1:])
                    else:
                        # Check if first line repeats the extra column header
                        if non_table_lines[0].strip().lower() == extra_col_header.strip().lower():
                            extra_col_values.extend(non_table_lines[1:])
                        else:
                            extra_col_values.extend(non_table_lines)
                elif len(non_table_lines) == data_rows_on_page - 1:
                    # All data lines, no header repeat on this page
                    extra_col_values.extend(non_table_lines)

    # --- Step 2: Re-attach extra column to the merged table ---
    if merged_tables and extra_col_values:
        table = merged_tables[0]
        # Add header for the extra column
        if extra_col_header and len(table) > 0:
            table[0].append(extra_col_header)

        # Add extra column values to data rows
        data_start = 1  # skip header
        for i, val in enumerate(extra_col_values):
            row_idx = data_start + i
            if row_idx < len(table):
                table[row_idx].append(val)
            else:
                break

        # Fill any remaining rows that didn't get an extra value
        expected_cols = len(table[0])
        for row in table:
            while len(row) < expected_cols:
                row.append("")

    # --- Step 3: Write to Excel ---
    wb = Workbook()
    ws = wb.active
    ws.title = "PDF Content"

    title_font = Font(name='Times New Roman', size=11, bold=True)
    data_font = Font(name='Times New Roman', size=11)
    alignment = Alignment(wrap_text=True, vertical='center')
    header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    current_row = 1

    # Write merged tables
    for table in merged_tables:
        if not table:
            continue

        for row_idx, row_data in enumerate(table):
            for col_idx, value in enumerate(row_data, start=1):
                cell_val = ""
                if value is not None:
                    cell_val = str(value).replace("\n", " ").strip() if isinstance(value, str) else value
                cell = ws.cell(row=current_row, column=col_idx, value=cell_val)
                cell.alignment = alignment
                if row_idx == 0:
                    cell.font = title_font
                    cell.fill = header_fill
                else:
                    cell.font = data_font
            current_row += 1

        current_row += 1  # Gap between tables

    # Auto-fit column widths
    for col_cells in ws.columns:
        max_length = 0
        col_letter = get_column_letter(col_cells[0].column)
        for cell in col_cells:
            try:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[col_letter].width = max(adjusted_width, 10)

    return wb


# ---------------------------------------------------------------------------
# EXTRACTION: Tables Only → Excel (tabula DataFrames)
# ---------------------------------------------------------------------------

def write_tables_to_excel(tables, excel_path=None):
    """
    Write tabula DataFrames to Excel.
    Merges tables with the same column count into a single sheet.
    """
    workbook = Workbook()

    title_font = Font(name='Times New Roman', size=11, bold=True)
    info_font = Font(name='Times New Roman', size=10)
    alignment = Alignment(wrap_text=True, vertical='center')
    header_fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')

    # Try to merge tables that have the same column structure
    merged_tables = []
    for table in tables:
        if merged_tables:
            last = merged_tables[-1]
            if list(last.columns) == list(table.columns):
                merged_tables[-1] = pd.concat([last, table], ignore_index=True)
                continue
        merged_tables.append(table.copy())

    for table_num, table in enumerate(merged_tables, start=1):
        sheet = workbook.create_sheet(title=f'Table_{table_num}')

        df = pd.DataFrame(table.values, columns=[str(col).title() if col else "" for col in table.columns])

        # Write headers
        for col_num, column_header in enumerate(df.columns, start=1):
            if column_header:
                cell = sheet.cell(row=1, column=col_num, value=column_header)
                cell.font = title_font
                cell.alignment = alignment
                cell.fill = header_fill
                sheet.column_dimensions[get_column_letter(col_num)].width = max(len(str(column_header)) + 2, 10)

        # Write data rows
        for row_num, (_, row) in enumerate(df.iterrows(), start=2):
            for col_num, value in enumerate(row, start=1):
                cell_val = str(value).replace("\n", " ").strip() if isinstance(value, str) else value
                cell = sheet.cell(row=row_num, column=col_num, value=cell_val)
                cell.font = info_font
                cell.alignment = alignment
                col_width = sheet.column_dimensions[get_column_letter(col_num)].width
                sheet.column_dimensions[get_column_letter(col_num)].width = max(col_width, min(len(str(value)) + 2, 50))

        # Adjust row heights
        for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, min_col=1, max_col=sheet.max_column):
            max_text_length = max((len(str(cell.value)) for cell in row if cell.value), default=0)
            sheet.row_dimensions[row[0].row].height = 35 + (max_text_length // 50) * 5

    workbook.remove(workbook.active)
    if excel_path:
        workbook.save(excel_path)
    return workbook


# ---------------------------------------------------------------------------
# HELPERS
# ---------------------------------------------------------------------------

def save_workbook(workbook, filename):
    excel_path = os.path.join(OUTPUT_FOLDER, filename)
    workbook.save(excel_path)
    return filename

def delete_files(*file_paths):
    for path in file_paths:
        if os.path.exists(path):
            try:
                os.remove(path)
            except PermissionError:
                pass


# ---------------------------------------------------------------------------
# ROUTES
# ---------------------------------------------------------------------------

@app.route('/')
def index():
    return render_template('pdftoexcel.html')

@app.route('/upload', methods=['POST'])
def upload():
    if 'pdfFile' not in request.files:
        return redirect(url_for('index'))

    file = request.files['pdfFile']

    if file.filename == '':
        return redirect(url_for('index'))

    if file and allowed_file(file.filename):
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        file.save(filepath)

        processing_option = request.form.get('processingOption')

        try:
            if processing_option == 'allText':
                pages_data = extract_pdf_content(filepath)
                excel_file = create_excel(pages_data)
                delete_files(filepath)
                out_name = 'output.xlsx'
                save_workbook(excel_file, out_name)
                return redirect(url_for('download_file', filename=out_name))
            elif processing_option == 'tablesOnly':
                tables = extract_tables_from_pdf(filepath)
                if tables:
                    excel_file = write_tables_to_excel(tables)
                    delete_files(filepath)
                    out_name = 'TablesOnly.xlsx'
                    save_workbook(excel_file, out_name)
                    return redirect(url_for('download_file', filename=out_name))
                else:
                    delete_files(filepath)
                    return 'No tables found in the PDF.'
            else:
                delete_files(filepath)
                return 'Invalid processing option'
        except Exception as e:
            delete_files(filepath)
            return f'Error processing PDF: {str(e)}', 500

    return 'Invalid file. Please upload a PDF file.'

@app.route('/download/<filename>')
def download_file(filename):
    safe_name = secure_filename(filename)
    file_path = os.path.join(OUTPUT_FOLDER, safe_name)
    if os.path.exists(file_path):
        return send_file(file_path, as_attachment=True)
    return 'File not found.', 404


if __name__ == "__main__":
    app.run(debug=True)